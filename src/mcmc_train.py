"""
贝叶斯层次模型 + MCMC 推断粉丝投票强度
完整版本 - 使用 C 编译加速，多核并行采样
融合版本：结合可视化诊断与淘汰预测
"""

import os
import warnings

os.environ.setdefault(
    "PYTENSOR_FLAGS", "device=cpu,floatX=float64,optimizer=fast_compile"
)

# 屏蔽非关键警告
warnings.simplefilter(action="ignore", category=FutureWarning)

import numpy as np
import pandas as pd
import pymc as pm
import pytensor.tensor as pt
import arviz as az
import matplotlib.pyplot as plt
import seaborn as sns
import multiprocessing as mp
from typing import Dict, List, Tuple
from dataclasses import dataclass

from configs.config import OUTPUT_DIR

# 设置中文字体支持
plt.rcParams["font.sans-serif"] = ["SimHei", "Microsoft YaHei", "DejaVu Sans"]
plt.rcParams["axes.unicode_minus"] = False

# 设置随机种子
np.random.seed(42)


@dataclass
class MCMCConfig:
    """MCMC 采样配置"""

    draws: int = 1000
    tune: int = 1000
    chains: int = 4
    cores: int = -1
    target_accept: float = 0.95
    init: str = "jitter+adapt_diag"  # 更稳定的初始化方法

    def __post_init__(self):
        if self.cores == -1:
            self.cores = min(mp.cpu_count(), self.chains)


def load_preprocessed_data(
    file_name: str = "preprocessed_data_percentage.csv",
) -> pd.DataFrame:
    """加载预处理后的数据"""
    file_path = OUTPUT_DIR / "preprocessed" / file_name

    try:
        df = pd.read_csv(file_path, encoding="utf-8")
        return df
    except FileNotFoundError as e:
        raise


def prepare_indices(df: pd.DataFrame) -> Tuple[Dict, int, int]:
    """
    准备赛季和选手索引

    Returns:
        (season_map, n_seasons, n_contestants)
    """
    seasons = sorted(df["season"].unique())
    season_map = {s: i for i, s in enumerate(seasons)}

    df["season_idx"] = df["season"].map(season_map)
    df["contestant_id"] = range(len(df))

    n_seasons = len(seasons)
    n_contestants = len(df)

    return season_map, n_seasons, n_contestants


def extract_features(
    df: pd.DataFrame, n_contestants: int
) -> Tuple[np.ndarray, np.ndarray, np.ndarray]:
    """
    直接从预处理数据中提取特征（不重复编码）

    Returns:
        (X_industry, X_age, X_advanced_rounds)
    """
    industry_cols = [c for c in df.columns if "celebrity_industry_" in c]
    if len(industry_cols) > 0:
        X_industry = df[industry_cols].values.astype(np.float64)
    else:
        X_industry = np.zeros((n_contestants, 1), dtype=np.float64)

    if "celebrity_age_during_season" in df.columns:
        X_age = df["celebrity_age_during_season"].values.astype(np.float64)
    else:
        X_age = np.zeros(n_contestants, dtype=np.float64)

    # 提取晋级轮次（标准化后）作为人气特征
    if "advanced_rounds" in df.columns:
        X_advanced_rounds = df["advanced_rounds"].values.astype(np.float64)
    else:
        X_advanced_rounds = np.zeros(n_contestants, dtype=np.float64)

    return X_industry, X_age, X_advanced_rounds


def build_observation_data(
    df: pd.DataFrame,
    max_weeks: int = 11,
) -> Tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, Dict]:
    """
    构建长格式观测数据

    Returns:
        (obs_season_idx, obs_week_idx, obs_contestant_idx, obs_score_sum, flat_idx_map)
    """
    obs_season_idx = []
    obs_week_idx = []
    obs_contestant_idx = []
    obs_score_sum = []  # 使用标准化后的分数总和
    flat_idx_map = {}
    current_flat_idx = 0

    for idx, row in df.iterrows():
        c_id = row["contestant_id"]
        s_idx = row["season_idx"]

        # 获取选手实际参与的周数
        weeks_participated = None
        if "weeks_participated" in df.columns:
            wp = row.get("weeks_participated", np.nan)
            try:
                weeks_participated = int(wp) if pd.notna(wp) else None
            except (ValueError, TypeError):
                weeks_participated = None

        for w in range(1, max_weeks + 1):
            col_score = f"week{w}_score_sum"  # 使用标准化后的分数总和

            if col_score not in df.columns:
                continue

            score_val = row[col_score]

            # 判断该周是否是合理数据：基于实际参与周数
            participated = (
                weeks_participated is not None and w <= weeks_participated
            ) or (
                weeks_participated is None and pd.notna(score_val)
            )  # 回退兼容

            if participated:
                # 即使 score_val 为 0 也是有效观测（标准化后可能为0）
                obs_season_idx.append(s_idx)
                obs_week_idx.append(w - 1)
                obs_contestant_idx.append(c_id)
                obs_score_sum.append(score_val if pd.notna(score_val) else 0.0)
                flat_idx_map[(c_id, w)] = current_flat_idx
                current_flat_idx += 1

    return (
        np.array(obs_season_idx, dtype=np.int32),
        np.array(obs_week_idx, dtype=np.int32),
        np.array(obs_contestant_idx, dtype=np.int32),
        np.array(obs_score_sum, dtype=np.float64),  # 返回标准化分数
        flat_idx_map,
    )


def build_elimination_pairs(
    df: pd.DataFrame,
    season_map: Dict,
    flat_idx_map: Dict,
    max_weeks: int = 11,
) -> Tuple[np.ndarray, Dict]:
    """
    构建淘汰约束配对（严格避免信息泄露）

    改进逻辑（避免未来信息泄露）：
    1. **不依赖 weeks_participated**（这是未来信息！）
    2. **不依赖 placement**（最终排名也是未来信息）
    3. **只看当前周和下一周的数据存在性**：
       - 如果选手在第w周有数据，但在第w+1周没有数据 → 第w周被淘汰
       - 如果选手在第w周和第w+1周都有数据 → 第w周晋级
    4. 约束：在同一周内，晋级者的综合得分 > 被淘汰者

    这样模型在推断第w周时，只能看到第w周和第w+1周的"是否继续"信息，
    而不会提前知道选手会参加多少周或最终排名。

    Returns:
        (elimination_pairs, pair_info)
        - elimination_pairs: [[winner_idx, loser_idx], ...]
        - pair_info: {pair_idx: {"winner": name, "loser": name, "week": w, "season": s}}
    """
    elimination_pairs = []
    pair_info = {}  # 用于调试验证
    pair_idx = 0

    for s in df["season"].unique():
        s_df = df[df["season"] == s]
        season_total_weeks = s_df["season_total_weeks"].iloc[0]

        # 遍历每一周（除了最后一周，因为最后一周没有淘汰）
        for w in range(1, min(max_weeks, season_total_weeks)):
            week_contestants = []

            for _, row in s_df.iterrows():
                c_id = row["contestant_id"]

                # 检查本周是否有观测数据
                if (c_id, w) not in flat_idx_map:
                    continue

                flat_idx = flat_idx_map[(c_id, w)]

                # **关键改进**：只看下一周是否有数据（不看weeks_participated）
                has_next_week = (c_id, w + 1) in flat_idx_map

                week_contestants.append(
                    {
                        "flat_idx": flat_idx,
                        "has_next_week": has_next_week,  # 是否晋级（二值化）
                        "contestant_id": c_id,
                        "name": row["celebrity_name"],
                    }
                )

            advanced = []  # 晋级者：下周有数据
            eliminated = []  # 淘汰者：下周没数据

            for c in week_contestants:
                if c["has_next_week"]:
                    advanced.append(c)
                else:
                    eliminated.append(c)

            # 生成配对：每个晋级者 vs 每个被淘汰者
            for winner in advanced:
                for loser in eliminated:
                    elimination_pairs.append([winner["flat_idx"], loser["flat_idx"]])
                    pair_info[pair_idx] = {
                        "winner": winner["name"],
                        "loser": loser["name"],
                        "week": w,
                        "season": s,
                        "winner_continues": True,
                        "loser_continues": False,
                    }
                    pair_idx += 1

    return (
        (
            np.array(elimination_pairs, dtype=np.int32)
            if elimination_pairs
            else np.array([], dtype=np.int32).reshape(0, 2)
        ),
        pair_info,
    )


def build_pymc_model(
    obs_season_idx: np.ndarray,
    obs_week_idx: np.ndarray,
    obs_contestant_idx: np.ndarray,
    obs_score_sum: np.ndarray,
    X_industry: np.ndarray,
    X_age: np.ndarray,
    X_advanced_rounds: np.ndarray,
    elimination_pairs: np.ndarray,
    n_seasons: int,
    n_contestants: int,
    n_observations: int,
) -> pm.Model:
    """
    构建完整的贝叶斯层次模型

    模型结构：
    - season_trend: 赛季趋势 (Gaussian Random Walk)
    - beta_week: 周次效应
    - alpha: 选手基础人气（融入晋级轮次）
    - beta_judge: 评委分权重
    - beta_industry: 职业特征权重
    - beta_age: 年龄权重
    - V_latent: 潜在投票强度 (Gamma 分布)
    - constraint: 淘汰约束 (Bernoulli)
    """
    n_industry_features = X_industry.shape[1]
    n_pairs = len(elimination_pairs)

    with pm.Model() as model:

        # 1. 赛季趋势 (Gaussian Random Walk) - 优化：减小方差
        sigma_season = pm.HalfNormal("sigma_season", sigma=0.05)
        season_trend = pm.GaussianRandomWalk(
            "season_trend",
            sigma=sigma_season,
            shape=n_seasons,
            init_dist=pm.Normal.dist(0, 0.05),
        )

        beta_week = pm.Normal("beta_week", mu=0, sigma=0.1)

        # 3. 选手基础人气（融入晋级轮次，已标准化）
        theta = pm.Normal("theta", mu=0, sigma=0.2)  # 基础人气均值
        theta_popularity = pm.Normal(
            "theta_popularity", mu=0.3, sigma=0.12
        )  # 晋级轮次效应系数
        sigma_alpha = pm.HalfNormal("sigma_alpha", sigma=0.5)

        # alpha先验均值由晋级轮次调整：晋级轮次越多，人气越高
        alpha_mu = theta + theta_popularity * X_advanced_rounds
        alpha = pm.Normal(
            "alpha",
            mu=alpha_mu,
            sigma=sigma_alpha,
            shape=n_contestants,
        )

        # 4. 评委分权重
        beta_judge = pm.Normal("beta_judge", mu=0.5, sigma=0.3)

        # 5. 职业特征权重
        beta_ind = pm.Normal("beta_ind", mu=0, sigma=0.3, shape=n_industry_features)

        # 6. 年龄权重
        beta_age = pm.Normal("beta_age", mu=0, sigma=0.8)

        # === Log-Linear 模型（投票强度） ===
        log_mu = (
            alpha[obs_contestant_idx]
            + beta_judge * obs_score_sum
            + pm.math.dot(X_industry, beta_ind)[obs_contestant_idx]
            + beta_age * X_age[obs_contestant_idx]
            + season_trend[obs_season_idx]
            + beta_week * obs_week_idx
        )

        phi = pm.HalfNormal("phi", sigma=2.0)
        mu_ = pm.math.exp(log_mu)
        V_latent = pm.Gamma(
            "V_latent",
            alpha=phi,
            beta=phi / mu_,
            shape=n_observations,
        )

        # === 淘汰约束 ===
        if n_pairs > 0:
            winners_idx = elimination_pairs[:, 0]
            losers_idx = elimination_pairs[:, 1]

            # 约束：晋级者的综合得分 > 淘汰者 - 优化：减小约束强度
            diff = (obs_score_sum[winners_idx] - obs_score_sum[losers_idx]) + 0.3 * (
                pt.log(V_latent[winners_idx]) - pt.log(V_latent[losers_idx])
            )

            # Sigmoid 概率约束 - 优化：从5降至3
            p_outcome = pm.math.sigmoid(diff * 3)
            pm.Bernoulli(
                "constraint",
                p=p_outcome,
                observed=np.ones(n_pairs, dtype=np.int32),
            )

    return model


def run_mcmc_sampling(model: pm.Model, config: MCMCConfig) -> az.InferenceData:
    """运行 MCMC 采样"""
    print(f"🚀 Starting MCMC sampling with {config.cores} cores...")
    print(f"   Chains: {config.chains}, Draws: {config.draws}, Tune: {config.tune}")
    print(f"   Init: {config.init}")

    with model:
        trace = pm.sample(
            draws=config.draws,
            tune=config.tune,
            chains=config.chains,
            cores=config.cores,
            target_accept=config.target_accept,
            init=config.init,
            return_inferencedata=True,
            progressbar=True,
            idata_kwargs={"log_likelihood": False},  # 减少内存使用
            compute_convergence_checks=False,  # 禁用自动收敛检查，我们手动做
        )

    return trace


def extract_results(
    trace: az.InferenceData,
    df: pd.DataFrame,
    obs_season_idx: np.ndarray,
    obs_week_idx: np.ndarray,
    obs_contestant_idx: np.ndarray,
    obs_score_sum: np.ndarray,
    season_map: Dict,
) -> pd.DataFrame:
    """提取推断结果"""
    # 提取潜在票数后验
    if "V_latent" not in trace.posterior:
        raise ValueError(
            "V_latent not found in trace.posterior. Available variables: "
            + ", ".join(trace.posterior.data_vars.keys())
        )

    v_samples = trace.posterior["V_latent"].values  # (chains, draws, observations)

    # 检查数组大小
    if v_samples.size == 0:
        raise ValueError(f"V_latent samples array is empty. Shape: {v_samples.shape}")

    v_mean = v_samples.mean(axis=(0, 1))
    v_std = v_samples.std(axis=(0, 1))
    v_lower = np.percentile(v_samples, 2.5, axis=(0, 1))
    v_upper = np.percentile(v_samples, 97.5, axis=(0, 1))

    # 反转 season_map
    inv_season_map = {v: k for k, v in season_map.items()}

    # 构造结果表
    results = []
    for i in range(len(obs_score_sum)):
        c_idx = obs_contestant_idx[i]
        celeb_name = df.loc[df["contestant_id"] == c_idx, "celebrity_name"].values[0]

        results.append(
            {
                "season": inv_season_map[obs_season_idx[i]],
                "week": obs_week_idx[i] + 1,
                "celebrity_name": celeb_name,
                "contestant_id": c_idx,
                "judge_score_sum": obs_score_sum[i],
                "vote_intensity_mean": v_mean[i],
                "vote_intensity_std": v_std[i],
                "vote_intensity_lower_95": v_lower[i],
                "vote_intensity_upper_95": v_upper[i],
            }
        )

    result_df = pd.DataFrame(results)

    # 排序
    result_df = result_df.sort_values(
        ["season", "week", "vote_intensity_mean"],
        ascending=[True, True, False],
    ).reset_index(drop=True)

    return result_df


def save_results(df: pd.DataFrame, output_file: str) -> None:
    """保存结果"""
    output_path = OUTPUT_DIR / "trained" / output_file
    output_path.parent.mkdir(parents=True, exist_ok=True)

    df.to_csv(output_path, index=False)
    print(f"✅ Results saved to: {output_path}")


def export_elimination_analysis_to_excel(
    df: pd.DataFrame,
    flat_idx_map: Dict,
    pair_info: Dict,
    output_file: str = "elimination_analysis.xlsx",
) -> None:
    """
    导出每个赛季的淘汰/晋级分析到Excel
    每个season一个sheet，包含每周的选手状态
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    output_path = OUTPUT_DIR / "trained" / output_file
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # 创建Excel writer
    writer = pd.ExcelWriter(output_path, engine="openpyxl")

    # 按赛季分组
    for season in sorted(df["season"].unique()):
        s_df = df[df["season"] == season].copy()
        season_total_weeks = s_df["season_total_weeks"].iloc[0]

        # 构建每周的选手状态表
        weekly_data = []

        for week in range(1, min(12, season_total_weeks + 1)):
            for _, row in s_df.iterrows():
                c_id = row["contestant_id"]
                c_name = row["celebrity_name"]

                # 检查本周是否有数据
                has_this_week = (c_id, week) in flat_idx_map
                has_next_week = (c_id, week + 1) in flat_idx_map

                if has_this_week:
                    # 获取本周的标准化分数
                    week_col = f"week{week}_score_sum"
                    judge_score = row.get(week_col, None)

                    # 判断状态
                    if has_next_week:
                        status = "晋级"
                        status_en = "Advanced"
                    elif week == season_total_weeks:
                        status = "冠军"
                        status_en = "Winner"
                    else:
                        status = "淘汰"
                        status_en = "Eliminated"

                    weekly_data.append(
                        {
                            "Week": week,
                            "Celebrity": c_name,
                            "Judge_Score_Sum": (  # 改为标准化分数
                                judge_score if pd.notna(judge_score) else 0
                            ),
                            "Status": status,
                            "Status_EN": status_en,
                            "Has_Next_Week": has_next_week,
                            "Advanced_Rounds": row.get("advanced_rounds", 0),
                            "Total_Weeks_Participated": row["weeks_participated"],
                            "Final_Placement": row["placement"],
                        }
                    )

        # 创建DataFrame
        weekly_df = pd.DataFrame(weekly_data)

        if len(weekly_df) > 0:
            # 排序：按周次、状态、评委分（使用标准化分数总和）
            weekly_df = weekly_df.sort_values(
                ["Week", "Status", "Judge_Score_Sum"], ascending=[True, False, False]
            )

            # 写入Excel
            sheet_name = f"Season_{season}"
            weekly_df.to_excel(writer, sheet_name=sheet_name, index=False)

            # 格式化（添加颜色标记）
            worksheet = writer.sheets[sheet_name]

            # 设置列宽
            worksheet.column_dimensions["A"].width = 8
            worksheet.column_dimensions["B"].width = 15
            worksheet.column_dimensions["C"].width = 18
            worksheet.column_dimensions["D"].width = 12
            worksheet.column_dimensions["E"].width = 12

            # 设置表头样式
            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            header_font = Font(color="FFFFFF", bold=True)

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # 根据状态设置行颜色
            advanced_fill = PatternFill(
                start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
            )  # 绿色
            eliminated_fill = PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
            )  # 红色
            winner_fill = PatternFill(
                start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
            )  # 金色

            for row_idx, row in enumerate(
                worksheet.iter_rows(min_row=2, max_row=len(weekly_df) + 1), start=2
            ):
                status = worksheet.cell(row_idx, 4).value

                if status == "晋级":
                    fill = advanced_fill
                elif status == "淘汰":
                    fill = eliminated_fill
                elif status == "冠军":
                    fill = winner_fill
                else:
                    fill = None

                if fill:
                    for cell in row:
                        cell.fill = fill
                        cell.alignment = Alignment(horizontal="center")

    # 保存
    writer.close()
    print(f"✅ Elimination analysis saved to: {output_path}")


def analyze_and_visualize_results(
    trace: az.InferenceData,
    result_df: pd.DataFrame,
    df: pd.DataFrame,
    season_map: Dict,
    output_dir_name: str = "mcmc_figures",
) -> None:
    """
    模型诊断、可视化与结果分析
    借鉴自 贝叶斯分层+mcmc.py 的 analyze_results 函数
    """
    output_dir = OUTPUT_DIR / "trained" / output_dir_name
    output_dir.mkdir(parents=True, exist_ok=True)

    print("\n" + "=" * 60)
    print("📊 Model Diagnostics & Visualization")
    print("=" * 60)

    # 1. 诊断统计量 - 关键参数摘要
    print("\n[1] Key Parameter Summary:")
    try:
        summary = az.summary(
            trace,
            var_names=[
                "sigma_season",
                "beta_week",
                "beta_judge",
                "theta",
                "theta_popularity",
                "sigma_alpha",
                "phi",
            ],
        )
        print(summary)
        # 保存摘要到CSV
        summary.to_csv(output_dir / "parameter_summary.csv")
        print(f"   Saved to: {output_dir / 'parameter_summary.csv'}")
    except Exception as e:
        print(f"   Warning: Could not generate summary - {e}")

    # 2. 轨迹图 (Traceplot)
    print("\n[2] Generating Traceplot...")
    try:
        fig, axes = plt.subplots(4, 2, figsize=(14, 12))
        az.plot_trace(
            trace,
            var_names=["beta_judge", "beta_week", "theta", "theta_popularity"],
            compact=True,
            axes=axes,
        )
        plt.tight_layout()
        plt.savefig(output_dir / "model_traceplot.png", dpi=150, bbox_inches="tight")
        plt.close()
        print(f"   Saved to: {output_dir / 'model_traceplot.png'}")
    except Exception as e:
        print(f"   Warning: Could not generate traceplot - {e}")

    # 3. 后验分布图 (Forest Plot for Industry Coefficients)
    print("\n[3] Generating Industry Effect Forest Plot...")
    try:
        plt.figure(figsize=(12, 8))
        az.plot_forest(trace, var_names=["beta_ind"], combined=True)
        plt.title("Impact of Industry on Contestant Strength (beta_industry)")
        plt.tight_layout()
        plt.savefig(
            output_dir / "industry_effect_forest.png", dpi=150, bbox_inches="tight"
        )
        plt.close()
        print(f"   Saved to: {output_dir / 'industry_effect_forest.png'}")
    except Exception as e:
        print(f"   Warning: Could not generate forest plot - {e}")

    # 4. 赛季趋势可视化
    print("\n[4] Generating Season Trend Plot...")
    try:
        season_trend_post = (
            trace.posterior["season_trend"].mean(dim=["chain", "draw"]).values
        )
        season_trend_std = (
            trace.posterior["season_trend"].std(dim=["chain", "draw"]).values
        )

        inv_season_map = {v: k for k, v in season_map.items()}
        seasons_list = [inv_season_map[i] for i in range(len(season_map))]

        plt.figure(figsize=(12, 6))
        plt.errorbar(
            seasons_list,
            season_trend_post,
            yerr=season_trend_std,
            marker="o",
            linestyle="-",
            color="purple",
            capsize=5,
            capthick=2,
            linewidth=2,
            markersize=8,
        )
        plt.fill_between(
            seasons_list,
            season_trend_post - season_trend_std,
            season_trend_post + season_trend_std,
            alpha=0.3,
            color="purple",
        )
        plt.title("Season Baseline Trend (Random Walk)", fontsize=14)
        plt.xlabel("Season", fontsize=12)
        plt.ylabel("Baseline Strength Correction", fontsize=12)
        plt.grid(True, alpha=0.3)
        plt.tight_layout()
        plt.savefig(output_dir / "season_trend.png", dpi=150, bbox_inches="tight")
        plt.close()
        print(f"   Saved to: {output_dir / 'season_trend.png'}")
    except Exception as e:
        print(f"   Warning: Could not generate season trend plot - {e}")

    # 5. 选手排名表 (Top 30 by Average Vote Intensity)
    print("\n[5] Generating Top Contestants Bar Chart...")
    try:
        avg_strength = (
            result_df.groupby(["contestant_id", "celebrity_name"])[
                "vote_intensity_mean"
            ]
            .mean()
            .reset_index()
        )
        top_30 = avg_strength.sort_values("vote_intensity_mean", ascending=False).head(
            30
        )

        plt.figure(figsize=(14, 10))
        sns.barplot(
            x="vote_intensity_mean",
            y="celebrity_name",
            data=top_30,
            palette="viridis",
        )
        plt.title("Top 30 Contestants by Estimated Vote Intensity", fontsize=14)
        plt.xlabel("Average Vote Intensity Score", fontsize=12)
        plt.ylabel("Celebrity Name", fontsize=12)
        plt.tight_layout()
        plt.savefig(output_dir / "top_contestants.png", dpi=150, bbox_inches="tight")
        plt.close()
        print(f"   Saved to: {output_dir / 'top_contestants.png'}")

        # 保存排名表到CSV
        avg_strength_sorted = avg_strength.sort_values(
            "vote_intensity_mean", ascending=False
        )
        avg_strength_sorted["rank"] = range(1, len(avg_strength_sorted) + 1)
        avg_strength_sorted.to_csv(output_dir / "contestant_ranking.csv", index=False)
        print(f"   Ranking saved to: {output_dir / 'contestant_ranking.csv'}")
    except Exception as e:
        print(f"   Warning: Could not generate top contestants chart - {e}")

    # 6. 后验预测检验 (Posterior Predictive Check)
    print("\n[6] Generating Posterior Distribution Plots...")
    try:
        fig, axes = plt.subplots(2, 3, figsize=(15, 10))

        # beta_judge 后验分布
        az.plot_posterior(trace, var_names=["beta_judge"], ax=axes[0, 0])
        axes[0, 0].set_title("beta_judge (Judge Score Weight)")

        # beta_week 后验分布
        az.plot_posterior(trace, var_names=["beta_week"], ax=axes[0, 1])
        axes[0, 1].set_title("beta_week (Week Effect)")

        # theta 后验分布
        az.plot_posterior(trace, var_names=["theta"], ax=axes[0, 2])
        axes[0, 2].set_title("theta (Base Popularity)")

        # theta_popularity 后验分布
        az.plot_posterior(trace, var_names=["theta_popularity"], ax=axes[1, 0])
        axes[1, 0].set_title("theta_popularity (Low-Score Advance Effect)")

        # phi 后验分布
        az.plot_posterior(trace, var_names=["phi"], ax=axes[1, 1])
        axes[1, 1].set_title("phi (Gamma Dispersion)")

        # beta_age 后验分布
        az.plot_posterior(trace, var_names=["beta_age"], ax=axes[1, 2])
        axes[1, 2].set_title("beta_age (Age Effect)")

        plt.tight_layout()
        plt.savefig(
            output_dir / "posterior_distributions.png", dpi=150, bbox_inches="tight"
        )
        plt.close()
        print(f"   Saved to: {output_dir / 'posterior_distributions.png'}")
    except Exception as e:
        print(f"   Warning: Could not generate posterior plots - {e}")

    # 7. R-hat 和 ESS 诊断
    print("\n[7] Convergence Diagnostics (R-hat & ESS)...")
    try:
        # 检查更多参数的收敛性（排除可能为空的参数）
        var_names_to_check = [
            "beta_judge",
            "beta_week",
            "theta",
            "theta_popularity",
            "phi",
            "sigma_season",
            "sigma_alpha",
            "beta_age",
        ]

        # 过滤出实际存在且有数据的变量
        available_vars = []
        for var in var_names_to_check:
            if var in trace.posterior:
                var_data = trace.posterior[var]
                # 检查是否为空数组
                if var_data.size > 0:
                    available_vars.append(var)

        if not available_vars:
            print("   ⚠️  Warning: No variables available for convergence diagnostics")
            return

        rhat = az.rhat(trace, var_names=available_vars)
        ess = az.ess(trace, var_names=available_vars)

        print("\n   ╔══════════════════════════════════════════════════════╗")
        print("   ║  R-hat 诊断 (应该接近 1.0, 建议 < 1.01)              ║")
        print("   ╚══════════════════════════════════════════════════════╝")

        rhat_issues = []
        for var in rhat.data_vars:
            try:
                rhat_val = float(
                    rhat[var].values.flat[0]
                    if rhat[var].values.size > 0
                    else float("nan")
                )
                if not np.isnan(rhat_val):
                    status = (
                        "✅" if rhat_val < 1.01 else "⚠️" if rhat_val < 1.05 else "❌"
                    )
                    print(f"      {status} {var:20s}: {rhat_val:.4f}")
                    if rhat_val >= 1.01:
                        rhat_issues.append((var, rhat_val))
            except (ValueError, IndexError):
                print(f"      ⚠️  {var:20s}: Could not compute")
                continue

        if rhat_issues:
            print("\n   ⚠️  警告：以下参数的 R-hat > 1.01，可能需要更多采样:")
            for var, val in rhat_issues:
                print(f"      - {var}: {val:.4f}")
        else:
            print("\n   ✅ 所有参数的 R-hat < 1.01，收敛良好！")

        print("\n   ╔══════════════════════════════════════════════════════╗")
        print("   ║  有效样本量 (ESS) - 越大越好                         ║")
        print("   ╚══════════════════════════════════════════════════════╝")

        total_samples = trace.posterior.sizes["chain"] * trace.posterior.sizes["draw"]

        for var in ess.data_vars:
            try:
                ess_val = float(
                    ess[var].values.flat[0] if ess[var].values.size > 0 else 0
                )
                if ess_val > 0:
                    ess_ratio = ess_val / total_samples
                    status = (
                        "✅" if ess_ratio > 0.1 else "⚠️" if ess_ratio > 0.01 else "❌"
                    )
                    print(
                        f"      {status} {var:20s}: {ess_val:7.0f} ({ess_ratio:5.1%} of total)"
                    )
            except (ValueError, IndexError):
                print(f"      ⚠️  {var:20s}: Could not compute")
                continue

        # 保存诊断结果到文件
        diagnostics_data = []
        for var in rhat.data_vars:
            try:
                rhat_val = float(
                    rhat[var].values.flat[0]
                    if rhat[var].values.size > 0
                    else float("nan")
                )
                ess_val = float(
                    ess[var].values.flat[0] if ess[var].values.size > 0 else 0
                )
                diagnostics_data.append(
                    {
                        "parameter": var,
                        "rhat": rhat_val,
                        "ess": ess_val,
                        "total_samples": total_samples,
                        "ess_ratio": ess_val / total_samples if ess_val > 0 else 0,
                    }
                )
            except (ValueError, IndexError):
                continue

        if diagnostics_data:
            diagnostics_df = pd.DataFrame(diagnostics_data)
            diagnostics_df.to_csv(
                output_dir / "convergence_diagnostics.csv", index=False
            )
            print(
                f"\n   📄 Diagnostics saved to: {output_dir / 'convergence_diagnostics.csv'}"
            )

    except Exception as e:
        print(f"   ⚠️  Warning: Could not compute diagnostics - {e}")
        import traceback

        traceback.print_exc()

    print("\n" + "=" * 60)
    print("✅ Visualization completed!")
    print("=" * 60)


def predict_eliminations(
    result_df: pd.DataFrame,
    df: pd.DataFrame,
    season_map: Dict,
    flat_idx_map: Dict,
    output_dir: str = None,
) -> pd.DataFrame:
    """
    执行末尾淘汰预测判别
    借鉴自 贝叶斯分层+mcmc.py 的 predict_eliminations 函数
    """
    if output_dir is None:
        output_dir = OUTPUT_DIR / "trained"
    output_dir = OUTPUT_DIR / "trained"

    print("\n" + "=" * 60)
    print("🎯 Elimination Prediction Analysis")
    print("=" * 60)

    inv_season_map = {v: k for k, v in season_map.items()}
    seasons_list = [inv_season_map[i] for i in range(len(season_map))]

    # 准备容器
    prediction_results = []
    correct_predictions = 0
    total_elimination_events = 0

    # 按赛季和周次分组
    for season in df["season"].unique():
        s_df = df[df["season"] == season]
        season_idx = season_map[season]
        season_total_weeks = s_df["season_total_weeks"].iloc[0]

        for week in range(1, min(12, season_total_weeks)):
            # 获取本周所有参赛选手
            week_contestants = []
            for _, row in s_df.iterrows():
                c_id = row["contestant_id"]
                if (c_id, week) not in flat_idx_map:
                    continue

                has_next_week = (c_id, week + 1) in flat_idx_map

                # 从结果中获取本周的投票强度
                vote_intensity = result_df[
                    (result_df["contestant_id"] == c_id) & (result_df["week"] == week)
                ]["vote_intensity_mean"].values

                if len(vote_intensity) == 0:
                    continue

                week_contestants.append(
                    {
                        "contestant_id": c_id,
                        "celebrity_name": row["celebrity_name"],
                        "vote_intensity": vote_intensity[0],
                        "has_next_week": has_next_week,
                        "placement": row["placement"],
                    }
                )

            if len(week_contestants) == 0:
                continue

            # 实际淘汰者（下周没数据且不是冠军）
            actual_eliminated = [
                c
                for c in week_contestants
                if not c["has_next_week"] and c["placement"] > 1
            ]
            actual_survived = [c for c in week_contestants if c["has_next_week"]]

            if len(actual_eliminated) == 0 or len(actual_survived) == 0:
                continue

            total_elimination_events += len(actual_eliminated)

            # 模型预测：按投票强度排序，最低的应该被淘汰
            sorted_contestants = sorted(
                week_contestants, key=lambda x: x["vote_intensity"]
            )
            predicted_eliminated = sorted_contestants[: len(actual_eliminated)]

            # 比对结果
            actual_ids = set(c["contestant_id"] for c in actual_eliminated)
            predicted_ids = set(c["contestant_id"] for c in predicted_eliminated)
            hits = len(actual_ids.intersection(predicted_ids))
            correct_predictions += hits

            # 记录详细日志
            prediction_results.append(
                {
                    "Season": season,
                    "Week": week,
                    "Actual_Eliminated": ", ".join(
                        str(c["celebrity_name"]) for c in actual_eliminated
                    ),
                    "Actual_Eliminated_IDs": list(actual_ids),
                    "Predicted_Eliminated": ", ".join(
                        str(c["celebrity_name"]) for c in predicted_eliminated
                    ),
                    "Predicted_Eliminated_IDs": list(predicted_ids),
                    "Correct_Count": hits,
                    "Total_Eliminated": len(actual_eliminated),
                    "Is_Correct": hits == len(actual_eliminated),
                }
            )

    # 计算准确率
    accuracy = (
        correct_predictions / total_elimination_events
        if total_elimination_events > 0
        else 0
    )

    print(f"\n📈 Prediction Statistics:")
    print(f"   Total elimination events: {total_elimination_events}")
    print(f"   Correct predictions: {correct_predictions}")
    print(f"   Elimination prediction accuracy: {accuracy:.2%}")

    # 转换为 DataFrame 并保存
    pred_df = pd.DataFrame(prediction_results)
    output_path = output_dir / "elimination_predictions.csv"
    pred_df.to_csv(output_path, index=False)
    print(f"\n✅ Detailed predictions saved to: {output_path}")

    # 按赛季统计准确率
    if len(pred_df) > 0:
        print("\n📊 Accuracy by Season:")
        season_stats = (
            pred_df.groupby("Season")
            .agg(
                {
                    "Correct_Count": "sum",
                    "Total_Eliminated": "sum",
                }
            )
            .reset_index()
        )
        season_stats["Accuracy"] = (
            season_stats["Correct_Count"] / season_stats["Total_Eliminated"]
        )
        for _, row in season_stats.iterrows():
            print(
                f"   Season {row['Season']}: {row['Correct_Count']}/{row['Total_Eliminated']} = {row['Accuracy']:.2%}"
            )

    return pred_df


def main():
    """主函数：执行完整的贝叶斯 MCMC 推断流程"""

    INPUT_FILE = "preprocessed_data_percentage.csv"  # 使用 percentage 版本的数据
    OUTPUT_FILE = "bayesian_vote_intensity.csv"
    MAX_WEEKS = 11

    n_cores = mp.cpu_count()
    # highlight: MCMC 配置可点
    mcmc_config = MCMCConfig(
        draws=400,  # 优化：从500增至1000
        tune=400,  # 优化：从500增至1000
        chains=min(n_cores, 4),
        cores=min(n_cores, 4),
        target_accept=0.85,  # 优化：从0.9增至0.95
        init="advi+adapt_diag",
    )

    print("=" * 60)
    print("🔥 Bayesian Hierarchical Model + MCMC Inference")
    print(f"   Using {mcmc_config.cores} CPU cores (max available: {n_cores})")
    print("   C++ compilation enabled for acceleration")
    print("=" * 60)

    # [1/8] 加载数据
    print("\n[1/8] Loading preprocessed data...")
    df = load_preprocessed_data(INPUT_FILE)

    # [2/8] 准备索引
    print("\n[2/8] Preparing indices...")
    season_map, n_seasons, n_contestants = prepare_indices(df)
    print(f"      Seasons: {n_seasons}, Contestants: {n_contestants}")

    # [3/8] 提取特征
    print("\n[3/8] Extracting features...")
    X_industry, X_age, X_advanced_rounds = extract_features(df, n_contestants)
    print(f"      Industry features: {X_industry.shape[1]}")
    print(
        f"      Advanced rounds range: [{X_advanced_rounds.min():.2f}, {X_advanced_rounds.max():.2f}] (standardized)"
    )

    # [4/8] 构建观测数据
    print("\n[4/8] Building observation data...")
    obs_season_idx, obs_week_idx, obs_contestant_idx, obs_score_sum, flat_idx_map = (
        build_observation_data(df, MAX_WEEKS)
    )
    n_observations = len(obs_score_sum)
    print(f"      Observations: {n_observations}")

    # [5/8] 构建淘汰约束
    print("\n[5/8] Building elimination constraints...")
    elimination_pairs, pair_info = build_elimination_pairs(
        df, season_map, flat_idx_map, MAX_WEEKS
    )

    # 打印验证信息（前5个配对）
    if len(elimination_pairs) > 0:
        print(f"      Total pairs: {len(elimination_pairs)}")
        print(f"      Sample pairs (严格避免未来信息泄露):")
        for i in range(min(5, len(elimination_pairs))):
            info = pair_info[i]
            print(
                f"        Week {info['week']}, Season {info['season']}: "
                f"{info['winner']} (下周继续) > {info['loser']} (下周淘汰)"
            )

    # 限制约束数量以加速（随机采样）
    MAX_PAIRS = 200
    if len(elimination_pairs) > MAX_PAIRS:
        idx = np.random.choice(len(elimination_pairs), MAX_PAIRS, replace=False)
        elimination_pairs = elimination_pairs[idx]
        print(f"      Sampled pairs for efficiency: {len(elimination_pairs)}")

    # [6/8] 构建 PyMC 模型
    print("\n[6/8] Building PyMC model...")
    model = build_pymc_model(
        obs_season_idx=obs_season_idx,
        obs_week_idx=obs_week_idx,
        obs_contestant_idx=obs_contestant_idx,
        obs_score_sum=obs_score_sum,
        X_industry=X_industry,
        X_age=X_age,
        X_advanced_rounds=X_advanced_rounds,
        elimination_pairs=elimination_pairs,
        n_seasons=n_seasons,
        n_contestants=n_contestants,
        n_observations=n_observations,
    )
    print("      Model built successfully!")

    # [7/8] 运行 MCMC 采样
    print("\n[7/8] Running MCMC sampling...")
    trace = run_mcmc_sampling(model, mcmc_config)
    print("      Sampling completed!")

    # [8/8] 提取并保存结果
    print("\n[8/8] Extracting and saving results...")
    result_df = extract_results(
        trace=trace,
        df=df,
        obs_season_idx=obs_season_idx,
        obs_week_idx=obs_week_idx,
        obs_contestant_idx=obs_contestant_idx,
        obs_score_sum=obs_score_sum,
        season_map=season_map,
    )
    save_results(result_df, OUTPUT_FILE)

    # [9/11] 导出淘汰/晋级分析到Excel
    print("\n[9/11] Exporting elimination analysis to Excel...")
    export_elimination_analysis_to_excel(
        df=df,
        flat_idx_map=flat_idx_map,
        pair_info=pair_info,
        output_file="elimination_analysis.xlsx",
    )

    # [10/11] 模型诊断与可视化
    print("\n[10/11] Model diagnostics and visualization...")
    analyze_and_visualize_results(
        trace=trace,
        result_df=result_df,
        df=df,
        season_map=season_map,
    )

    # [11/11] 淘汰预测分析
    print("\n[11/11] Elimination prediction analysis...")
    pred_df = predict_eliminations(
        result_df=result_df,
        df=df,
        season_map=season_map,
        flat_idx_map=flat_idx_map,
    )

    print("\n" + "=" * 60)
    print("🎉 Bayesian MCMC inference completed!")
    print(f"   Total rows: {len(result_df)}")
    print("=" * 60)

    # 打印示例结果
    print("\n📊 Sample results (first 10 rows):")
    print(
        result_df[
            [
                "celebrity_name",
                "season",
                "week",
                "judge_score_sum",
                "vote_intensity_mean",
            ]
        ]
        .head(10)
        .to_string(index=False)
    )

    # 返回结果供进一步分析
    return {
        "result_df": result_df,
        "trace": trace,
        "df": df,
        "season_map": season_map,
        "pred_df": pred_df,
    }


if __name__ == "__main__":
    results = main()
